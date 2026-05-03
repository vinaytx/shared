"""
PDF Section Comparator
======================
Reads the output directories produced by pdf_section_extractor.py, matches
sections by heading, and produces a side-by-side comparison that includes
every section from both documents — with a clearly marked blank placeholder
wherever a section is missing from one of the documents.

Outputs (in output_dir/):
    _COMPARISON_REPORT.html   Interactive side-by-side web report
    _COMPARISON_REPORT.xlsx   Excel workbook (4 sheets)
    _MATCHED.json             Machine-readable results

Usage:
    python pdf_compare.py <dir_a> <dir_b> [output_dir] [options]

    dir_a / dir_b  — directories created by:
                     python pdf_section_extractor.py your.pdf <dir>/

Options:
    --text-model MODEL      Ollama text model for AI summaries (default: llama3.2)
    --ollama-url URL        Ollama server URL (default: http://localhost:11434)
    --skip-summary          Skip AI summaries (diff only, much faster)
    --fuzzy-threshold N     Min title similarity 0-100 to count as a match (default: 75)
    --report-format         html | text | both  (default: html)

Requirements:
    pip install requests openpyxl
    pdf_section_extractor.py must be in the same directory
"""

import os
import re
import sys
import json
import argparse
import difflib
import textwrap
import time
from pathlib import Path
from datetime import datetime

# ── Import shared loaders and constants from the extractor ───────────────────
_HERE = Path(globals().get("__file__", ".")).parent
sys.path.insert(0, str(_HERE))

try:
    from pdf_section_extractor import (
        load_sections_json,
        load_image_catalog,
        load_checkpoint,
        step_is_done,
        DEFAULT_OLLAMA_URL,
        SECTIONS_JSON_FILE,
        IMAGE_CATALOG_FILE,
        PIPELINE_STATE_FILE,
        STEP_WRITE,
    )
except ImportError:
    print("✗  Could not import pdf_section_extractor.py")
    print("   Ensure it is in the same directory as this script.")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════

DEFAULT_TEXT_MODEL  = "gemma4:31b-cloud"
SUMMARY_MAX_CHARS   = 4000

SUMMARY_SYSTEM = (
    "You are a precise technical document analyst. "
    "Compare two versions of the same document section and write a clear, "
    "concise plain-prose summary of the differences. "
    "Focus on: added content, removed content, changed facts or figures, "
    "reworded claims, and structural changes. "
    "Be specific — mention actual words, numbers, or phrases when relevant. "
    "If the sections are identical say so in one sentence. "
    "No bullet points, no markdown."
)

SUMMARY_TEMPLATE = """\
Section: {title}

=== VERSION A ({name_a}) ===
{text_a}

=== VERSION B ({name_b}) ===
{text_b}

Summarise the differences between Version A and Version B.\
"""


# ══════════════════════════════════════════════════════════════════
# LOAD EXTRACTION DIRECTORY
# ══════════════════════════════════════════════════════════════════

def load_extraction_dir(directory: str, label: str = "") -> tuple:
    """
    Validate and load everything from a pdf_section_extractor.py output dir.

    Checks:
      1. Directory exists
      2. _sections.json is present
      3. Pipeline completed (warns if not, continues with partial data)
      4. Loads sections, image catalog, and metadata

    Returns:
        (sections, image_catalog, meta, pipeline_state)
    """
    tag = f" [{label}]" if label else ""
    d   = Path(directory)

    if not d.exists():
        print(f"✗{tag}  Directory not found: {directory}")
        sys.exit(1)

    if not (d / SECTIONS_JSON_FILE).exists():
        print(f"✗{tag}  No {SECTIONS_JSON_FILE} in '{directory}'.")
        print(f"   Run the extractor first:")
        print(f"   python pdf_section_extractor.py your_file.pdf {directory}/")
        sys.exit(1)

    # Pipeline state check
    pipeline_state = {}
    if (d / PIPELINE_STATE_FILE).exists():
        pipeline_state = load_checkpoint(d)
        if not step_is_done(d, STEP_WRITE):
            completed = list(pipeline_state.get("steps", {}).keys())
            print(f"⚠{tag}  Pipeline incomplete in '{directory}'.")
            print(f"   Completed: {completed or 'none'}")
            print(f"   Resume  : python pdf_section_extractor.py <pdf> {directory}/ --resume")
            print(f"   Continuing with partial data...\n")
    else:
        print(f"  ⚠{tag}  No {PIPELINE_STATE_FILE} — cannot verify pipeline completion.\n")

    # Load sections
    try:
        sections, ocr_stats, meta = load_sections_json(d)
    except FileNotFoundError as exc:
        print(f"✗{tag}  {exc}")
        sys.exit(1)
    except Exception as exc:
        print(f"✗{tag}  Failed to load {SECTIONS_JSON_FILE}: {exc}")
        sys.exit(1)

    if not sections:
        print(f"✗{tag}  No sections found in '{directory}'. Re-run the extractor.")
        sys.exit(1)

    # Load image catalog (optional)
    image_catalog = []
    if (d / IMAGE_CATALOG_FILE).exists():
        try:
            image_catalog = load_image_catalog(d).get("images", [])
        except Exception as exc:
            print(f"  ⚠{tag}  Could not load {IMAGE_CATALOG_FILE}: {exc}")

    return sections, image_catalog, meta, pipeline_state


def _dir_summary_str(label: str, directory: str, sections: list,
                     image_catalog: list, meta: dict) -> str:
    """Return a multi-line loading summary string for one directory."""
    source = meta.get("source_pdf") or Path(directory).name
    counts = {}
    for s in sections:
        counts[s.level] = counts.get(s.level, 0) + 1
    imgs_with_text = sum(1 for i in image_catalog
                         if i.get("ocr_text", "").strip()
                         and i["ocr_text"].strip() != "[NO TEXT]")
    imgs_described = sum(1 for i in image_catalog if i.get("description", "").strip())

    lines = [
        f"  ✓  [{label}]  {source}",
        f"       Extracted : {meta.get('extracted_at', 'unknown')}",
        f"       Sections  : {len(sections)}  "
        f"({counts.get('chapter',0)} chapters, "
        f"{counts.get('section',0)} sections, "
        f"{counts.get('subsection',0)} subsections)",
        f"       Images    : {len(image_catalog)} saved  |  "
        f"{imgs_with_text} with OCR text  |  "
        f"{imgs_described} with AI descriptions",
    ]
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════
# TITLE NORMALISATION AND SECTION MATCHING
# ══════════════════════════════════════════════════════════════════

def normalise_title(title: str) -> str:
    """Strip leading numbers / punctuation and lowercase for fuzzy matching."""
    t = title.lower().strip()
    t = re.sub(r"^[\d\.\:\-\s]+", "", t)
    t = re.sub(r"[^\w\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def title_similarity(a: str, b: str) -> float:
    """Return 0-100 similarity score between two normalised titles."""
    na, nb = normalise_title(a), normalise_title(b)
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio() * 100


def match_sections(sections_a: list, sections_b: list, threshold: float = 75.0):
    """
    Greedily match sections from A to sections in B by title similarity.
    Only sections with the same structural level are candidates.

    Returns:
        matched   — list of (section_a, section_b, score)
        only_in_a — sections from A with no match in B
        only_in_b — sections from B with no match in A
    """
    used_b  = set()
    matched = []

    for sec_a in sections_a:
        best_score, best_b, best_idx = 0.0, None, -1
        for idx, sec_b in enumerate(sections_b):
            if idx in used_b or sec_a.level != sec_b.level:
                continue
            score = title_similarity(sec_a.title, sec_b.title)
            if score > best_score:
                best_score, best_b, best_idx = score, sec_b, idx
        if best_b is not None and best_score >= threshold:
            matched.append((sec_a, best_b, best_score))
            used_b.add(best_idx)

    matched_a = {id(m[0]) for m in matched}
    matched_b = {id(m[1]) for m in matched}
    only_in_a = [s for s in sections_a if id(s) not in matched_a]
    only_in_b = [s for s in sections_b if id(s) not in matched_b]

    return matched, only_in_a, only_in_b


# ══════════════════════════════════════════════════════════════════
# DIFF HELPERS
# ══════════════════════════════════════════════════════════════════

def text_diff_stats(text_a: str, text_b: str) -> dict:
    """Return line-level and word-level change statistics."""
    la, lb = text_a.splitlines(), text_b.splitlines()
    wa, wb = text_a.split(), text_b.split()
    opcodes = difflib.SequenceMatcher(None, la, lb).get_opcodes()
    return {
        "line_similarity_pct": round(
            difflib.SequenceMatcher(None, la, lb).ratio() * 100, 1),
        "word_similarity_pct": round(
            difflib.SequenceMatcher(None, wa, wb).ratio() * 100, 1),
        "lines_added":     sum(j2-j1 for t,i1,i2,j1,j2 in opcodes if t in ("insert","replace")),
        "lines_removed":   sum(i2-i1 for t,i1,i2,j1,j2 in opcodes if t in ("delete","replace")),
        "lines_unchanged": sum(i2-i1 for t,i1,i2,j1,j2 in opcodes if t == "equal"),
        "total_lines_a": len(la),
        "total_lines_b": len(lb),
        "total_words_a": len(wa),
        "total_words_b": len(wb),
    }


def unified_diff(text_a: str, text_b: str, label_a: str, label_b: str) -> str:
    diff = difflib.unified_diff(
        text_a.splitlines(keepends=True),
        text_b.splitlines(keepends=True),
        fromfile=label_a, tofile=label_b, lineterm=""
    )
    return "\n".join(diff)


def side_by_side_diff(text_a: str, text_b: str, width: int = 80) -> list:
    """Return list of (left, right, tag) tuples for HTML rendering."""
    la, lb = text_a.splitlines(), text_b.splitlines()
    rows   = []

    def pad(s):
        return s[:width].ljust(width) if len(s) <= width else s[:width-1] + "…"

    for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(None, la, lb).get_opcodes():
        if tag == "equal":
            for a, b in zip(la[i1:i2], lb[j1:j2]):
                rows.append((pad(a), pad(b), "equal"))
        elif tag == "replace":
            ba, bb = la[i1:i2], lb[j1:j2]
            for a, b in zip(ba, bb):
                rows.append((pad(a), pad(b), "replace"))
            for a in ba[len(bb):]:
                rows.append((pad(a), " "*width, "delete"))
            for b in bb[len(ba):]:
                rows.append((" "*width, pad(b), "insert"))
        elif tag == "delete":
            for a in la[i1:i2]:
                rows.append((pad(a), " "*width, "delete"))
        elif tag == "insert":
            for b in lb[j1:j2]:
                rows.append((" "*width, pad(b), "insert"))
    return rows


# ══════════════════════════════════════════════════════════════════
# OLLAMA SUMMARISATION
# ══════════════════════════════════════════════════════════════════

def check_ollama(base_url: str, model: str) -> bool:
    import requests as _r
    try:
        resp = _r.get(f"{base_url}/api/tags", timeout=5)
        resp.raise_for_status()
        available = [m["name"].split(":")[0] for m in resp.json().get("models", [])]
        if model.split(":")[0] not in available:
            print(f"  ⚠  Model '{model}' not found. Available: {', '.join(available)}")
            print(f"     Run: ollama pull {model}")
            return False
        return True
    except Exception as exc:
        print(f"  ⚠  Ollama unavailable: {exc}")
        return False


def _call_ollama(base_url: str, model: str, prompt: str, retries: int = 2) -> str:
    import requests as _r
    payload = {
        "model": model, "system": SUMMARY_SYSTEM,
        "prompt": prompt, "stream": False,
        "options": {"temperature": 0.2, "num_predict": -1},
    }
    for attempt in range(retries + 1):
        try:
            resp = _r.post(f"{base_url}/api/generate", json=payload, timeout=180)
            resp.raise_for_status()
            return resp.json().get("response", "").strip()
        except Exception as exc:
            if attempt < retries:
                time.sleep(2)
            else:
                return f"[Error: {exc}]"
    return "[Unavailable]"


def summarise_all_matches(matched: list, name_a: str, name_b: str,
                          base_url: str, model: str,
                          skip_summary: bool = False) -> list:
    """
    Return extended list of (sec_a, sec_b, score, summary) tuples.
    Identical sections get a canned line; changed sections go to Ollama.
    """
    if skip_summary:
        return [(sa, sb, sc, "") for sa, sb, sc in matched]

    if not check_ollama(base_url, model):
        msg = f"[LLM unavailable — run: ollama pull {model}]"
        return [(sa, sb, sc, msg) for sa, sb, sc in matched]

    print(f"  Using '{model}' to summarise {len(matched)} section(s)...\n")
    results = []
    for i, (sa, sb, score) in enumerate(matched, 1):
        sim = text_diff_stats(sa.content(), sb.content())["line_similarity_pct"]
        print(f"  [{i}/{len(matched)}] {sa.title}  ({sim}% similar)", end=" ... ", flush=True)
        if sim == 100.0:
            summary = "The two versions of this section are identical."
            print("identical")
        else:
            ta = sa.content()[:SUMMARY_MAX_CHARS]
            tb = sb.content()[:SUMMARY_MAX_CHARS]
            prompt = SUMMARY_TEMPLATE.format(
                title=sa.title, name_a=name_a, name_b=name_b,
                text_a=ta, text_b=tb)
            summary = _call_ollama(base_url, model, prompt)
            preview = summary[:70].replace("\n", " ")
            print(f'✓ "{preview}{"..." if len(summary) > 70 else ""}"')
        results.append((sa, sb, score, summary))
    return results


# ══════════════════════════════════════════════════════════════════
# BUILD UNIFIED ROW LIST
# ══════════════════════════════════════════════════════════════════

MISSING_PLACEHOLDER = "— not present —"

def build_all_rows(matched: list, only_a: list, only_b: list) -> list:
    """
    Produce a single ordered list of dicts covering EVERY section from
    both documents.  Missing sections carry a MISSING_PLACEHOLDER so
    every output format can render a blank cell alongside real content.

    Row dict keys:
        status          "CHANGED" | "IDENTICAL" | "ONLY IN A" | "ONLY IN B"
        level           structural level string
        title_a         title from A  (MISSING_PLACEHOLDER if absent)
        title_b         title from B  (MISSING_PLACEHOLDER if absent)
        content_a       full text from A  (empty string if absent)
        content_b       full text from B  (empty string if absent)
        title_score     float 0-100  (0 if unmatched)
        stats           text_diff_stats dict  (all zeros if unmatched)
        summary         AI summary string  (empty if unmatched or skipped)
        sec_a           Section object or None
        sec_b           Section object or None
    """
    rows = []

    # Matched pairs
    for item in matched:
        sa, sb, score, summary = item if len(item) == 4 else (*item, "")
        stats  = text_diff_stats(sa.content(), sb.content())
        status = "IDENTICAL" if stats["line_similarity_pct"] == 100.0 else "CHANGED"
        rows.append({
            "status":      status,
            "level":       sa.level,
            "title_a":     sa.title,
            "title_b":     sb.title,
            "content_a":   sa.content(),
            "content_b":   sb.content(),
            "title_score": round(score, 1),
            "stats":       stats,
            "summary":     summary,
            "sec_a":       sa,
            "sec_b":       sb,
        })

    # Sections only in A
    zero_stats = {k: 0 for k in text_diff_stats("", "").keys()}
    for sec in only_a:
        rows.append({
            "status":      "ONLY IN A",
            "level":       sec.level,
            "title_a":     sec.title,
            "title_b":     MISSING_PLACEHOLDER,
            "content_a":   sec.content(),
            "content_b":   "",
            "title_score": 0.0,
            "stats":       dict(zero_stats),
            "summary":     "",
            "sec_a":       sec,
            "sec_b":       None,
        })

    # Sections only in B
    for sec in only_b:
        rows.append({
            "status":      "ONLY IN B",
            "level":       sec.level,
            "title_a":     MISSING_PLACEHOLDER,
            "title_b":     sec.title,
            "content_a":   "",
            "content_b":   sec.content(),
            "title_score": 0.0,
            "stats":       dict(zero_stats),
            "summary":     "",
            "sec_a":       None,
            "sec_b":       sec,
        })

    return rows


# ══════════════════════════════════════════════════════════════════
# HTML REPORT
# ══════════════════════════════════════════════════════════════════

def _sim_colour(pct: float) -> str:
    if pct >= 90: return "#22c55e"
    if pct >= 70: return "#eab308"
    if pct >= 40: return "#f97316"
    return "#ef4444"


def write_html_report(rows: list, name_a: str, name_b: str,
                      out_dir: Path,
                      meta_a: dict = None, meta_b: dict = None) -> Path:
    """Write a self-contained HTML comparison report."""

    def esc(s: str) -> str:
        return (str(s)
                .replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    def diff_table_body(text_a: str, text_b: str) -> str:
        colours = {
            "equal":   ("", ""),
            "replace": ("#fef3c7", "#dbeafe"),
            "delete":  ("#fee2e2", ""),
            "insert":  ("", "#dbeafe"),
        }
        html_rows = []
        for left, right, tag in side_by_side_diff(text_a, text_b):
            cl, cr = colours[tag]
            sl = f' style="background:{cl}"' if cl else ""
            sr = f' style="background:{cr}"' if cr else ""
            mk = {"equal":" ","replace":"~","delete":"-","insert":"+"}[tag]
            html_rows.append(
                f'<tr><td class="ln">{mk}</td>'
                f'<td{sl}><pre>{esc(left)}</pre></td>'
                f'<td{sr}><pre>{esc(right)}</pre></td></tr>'
            )
        return "\n".join(html_rows)

    # ── Build section cards ───────────────────────────────────────
    cards = []
    matched_count  = sum(1 for r in rows if r["status"] in ("CHANGED","IDENTICAL"))
    only_a_count   = sum(1 for r in rows if r["status"] == "ONLY IN A")
    only_b_count   = sum(1 for r in rows if r["status"] == "ONLY IN B")
    sims = [r["stats"]["line_similarity_pct"]
            for r in rows if r["status"] in ("CHANGED","IDENTICAL")]
    avg_sim = round(sum(sims)/len(sims),1) if sims else 0

    for i, row in enumerate(rows, 1):
        status   = row["status"]
        sim      = row["stats"]["line_similarity_pct"]
        colour   = _sim_colour(sim) if status in ("CHANGED","IDENTICAL") else "#9ca3af"
        anchor   = f"sec{i}"
        is_missing_a = (status == "ONLY IN B")
        is_missing_b = (status == "ONLY IN A")

        # Status badge colour
        badge_colours = {
            "IDENTICAL": ("065f46","a7f3d0"),
            "CHANGED":   ("92400e","fde68a"),
            "ONLY IN A": ("1e40af","bfdbfe"),
            "ONLY IN B": ("065f46","bbf7d0"),
        }
        bc_bg, bc_fg = badge_colours.get(status, ("374151","d1d5db"))

        # Stat pills
        if status in ("CHANGED","IDENTICAL"):
            stat_pills = (
                f'<span>Lines: <b>{row["stats"]["total_lines_a"]}</b>'
                f' → <b>{row["stats"]["total_lines_b"]}</b></span>'
                f'<span class="added">+{row["stats"]["lines_added"]}</span>'
                f'<span class="removed">-{row["stats"]["lines_removed"]}</span>'
                f'<span class="unchanged">={row["stats"]["lines_unchanged"]}</span>'
            )
        else:
            stat_pills = ""

        # AI summary
        summary_html = ""
        if row.get("summary") and row["summary"].strip() and not row["summary"].startswith("["):
            summary_html = f"""
          <div class="ai-summary">
            <div class="ai-label">✦ AI Summary</div>
            <p>{esc(row['summary'])}</p>
          </div>"""

        # Diff section
        if status == "IDENTICAL":
            diff_html = '<div class="identical-note">✓ Sections are identical</div>'
        elif is_missing_a:
            diff_html = f"""
          <div class="missing-pane missing-a">
            <div class="missing-label">Not present in {esc(name_a)}</div>
          </div>
          <div class="content-pane">
            <div class="pane-label">{esc(name_b)}</div>
            <pre>{esc(row['content_b'])}</pre>
          </div>"""
        elif is_missing_b:
            diff_html = f"""
          <div class="content-pane">
            <div class="pane-label">{esc(name_a)}</div>
            <pre>{esc(row['content_a'])}</pre>
          </div>
          <div class="missing-pane missing-b">
            <div class="missing-label">Not present in {esc(name_b)}</div>
          </div>"""
        else:
            diff_tb = diff_table_body(row["content_a"], row["content_b"])
            diff_html = f"""
          <details {'open' if sim < 100 else ''}>
            <summary>Show diff</summary>
            <div class="diff-scroll">
              <table class="diff-table">
                <thead><tr>
                  <th class="ln"></th>
                  <th>{esc(name_a)}</th>
                  <th>{esc(name_b)}</th>
                </tr></thead>
                <tbody>{diff_tb}</tbody>
              </table>
            </div>
          </details>"""

        cards.append(f"""
      <div class="card" id="{anchor}">
        <div class="card-header">
          <div class="card-left">
            <span class="badge" style="background:#{bc_bg};color:#{bc_fg}">{esc(status)}</span>
            <span class="level-badge">{esc(row['level'].upper())}</span>
            <span class="card-title">{esc(row['title_a'] if not is_missing_a else row['title_b'])}</span>
          </div>
          <div class="card-right">
            {'<span class="sim-score" style="color:'+colour+'">'+str(sim)+'% similar</span>'
             if status in ("CHANGED","IDENTICAL") else ""}
          </div>
        </div>
        {'<div class="stat-row">'+stat_pills+'</div>' if stat_pills else ''}
        {summary_html}
        <div class="diff-area">{diff_html}</div>
      </div>""")

    # ── TOC ───────────────────────────────────────────────────────
    toc_items = "".join(
        f'<li><a href="#sec{i}">'
        f'<span class="toc-status toc-{row["status"].lower().replace(" ","-")}">'
        f'{"●"}</span> '
        f'{esc(row["title_a"] if row["status"] != "ONLY IN B" else row["title_b"])}'
        f'</a></li>'
        for i, row in enumerate(rows, 1)
    )

    meta_a = meta_a or {}
    meta_b = meta_b or {}

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PDF Comparison — {esc(name_a)} vs {esc(name_b)}</title>
<style>
  :root{{
    --bg:#0f172a;--surface:#1e293b;--border:#334155;
    --text:#e2e8f0;--muted:#94a3b8;--accent:#38bdf8;
    --green:#22c55e;--yellow:#eab308;--red:#ef4444;--orange:#f97316;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Courier New',monospace;background:var(--bg);color:var(--text);line-height:1.6}}
  a{{color:var(--accent);text-decoration:none}}
  a:hover{{text-decoration:underline}}

  .layout{{display:flex;min-height:100vh}}
  .sidebar{{width:260px;background:var(--surface);border-right:1px solid var(--border);
    padding:20px 14px;position:sticky;top:0;height:100vh;overflow-y:auto;flex-shrink:0}}
  .sidebar h2{{font-size:10px;text-transform:uppercase;letter-spacing:2px;
    color:var(--muted);margin-bottom:12px}}
  .sidebar ul{{list-style:none}}
  .sidebar li{{font-size:11px;margin-bottom:5px;line-height:1.4}}
  .sidebar li a{{color:var(--text)}}
  .sidebar li a:hover{{color:var(--accent)}}
  .toc-status{{font-size:8px;margin-right:3px}}
  .toc-changed{{color:#eab308}}
  .toc-identical{{color:#22c55e}}
  .toc-only-in-a{{color:#60a5fa}}
  .toc-only-in-b{{color:#34d399}}

  .main{{flex:1;padding:28px;max-width:1400px;overflow:hidden}}

  .page-header{{margin-bottom:32px;padding-bottom:20px;border-bottom:1px solid var(--border)}}
  .page-header h1{{font-size:26px;font-weight:700;color:var(--accent);margin-bottom:16px}}
  .meta-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px}}
  .meta-box{{background:var(--surface);border:1px solid var(--border);
    border-radius:8px;padding:12px 16px}}
  .meta-box .lbl{{font-size:10px;text-transform:uppercase;letter-spacing:1.5px;
    color:var(--muted)}}
  .meta-box .val{{font-size:18px;font-weight:700;margin-top:3px}}
  .meta-box .sub{{font-size:10px;color:var(--muted);margin-top:2px;word-break:break-all}}

  .section-heading{{font-size:13px;text-transform:uppercase;letter-spacing:1px;
    color:var(--muted);font-weight:700;margin:24px 0 12px}}

  .card{{background:var(--surface);border:1px solid var(--border);
    border-radius:8px;margin-bottom:16px;overflow:hidden}}
  .card-header{{display:flex;justify-content:space-between;align-items:center;
    padding:14px 18px;border-bottom:1px solid var(--border);gap:10px;flex-wrap:wrap}}
  .card-left{{display:flex;align-items:center;gap:8px;flex:1;min-width:0}}
  .card-title{{font-size:14px;font-weight:600;white-space:nowrap;
    overflow:hidden;text-overflow:ellipsis}}
  .card-right{{flex-shrink:0}}
  .sim-score{{font-size:13px;font-weight:700}}
  .badge{{font-size:10px;font-weight:700;padding:2px 7px;border-radius:4px;
    text-transform:uppercase;letter-spacing:1px;flex-shrink:0}}
  .level-badge{{font-size:9px;color:var(--muted);border:1px solid var(--border);
    padding:1px 5px;border-radius:3px;flex-shrink:0}}
  .stat-row{{display:flex;gap:14px;padding:8px 18px;font-size:11px;
    color:var(--muted);border-bottom:1px solid var(--border);flex-wrap:wrap}}
  .added{{color:#22c55e}} .removed{{color:#ef4444}} .unchanged{{color:var(--muted)}}

  .ai-summary{{padding:12px 18px;background:#0c1a2e;border-bottom:1px solid var(--border)}}
  .ai-label{{font-size:10px;text-transform:uppercase;letter-spacing:2px;
    color:var(--accent);margin-bottom:6px}}
  .ai-summary p{{font-size:12px;line-height:1.7;color:#cbd5e1;font-family:Georgia,serif}}

  .diff-area{{padding:0}}
  .identical-note{{padding:12px 18px;font-size:12px;color:#22c55e}}

  .content-pane{{padding:12px 18px}}
  .pane-label{{font-size:10px;text-transform:uppercase;letter-spacing:1px;
    color:var(--muted);margin-bottom:6px}}
  .content-pane pre{{font-size:11px;white-space:pre-wrap;word-break:break-word;
    color:var(--text);line-height:1.6}}

  .missing-pane{{padding:20px 18px;display:flex;align-items:center;justify-content:center}}
  .missing-pane.missing-a{{background:#1a1a2e;border-right:1px solid var(--border)}}
  .missing-pane.missing-b{{background:#1a2e1a}}
  .missing-label{{font-size:12px;color:#6b7280;font-style:italic}}

  details summary{{cursor:pointer;padding:10px 18px;font-size:11px;
    color:var(--accent);user-select:none}}
  details summary:hover{{background:rgba(56,189,248,.05)}}
  .diff-scroll{{overflow-x:auto}}
  .diff-table{{width:100%;border-collapse:collapse;font-size:11px}}
  .diff-table th{{padding:6px 10px;text-align:left;background:var(--bg);
    color:var(--muted);font-size:10px;text-transform:uppercase;
    border-bottom:1px solid var(--border);position:sticky;top:0}}
  .diff-table td{{padding:1px 10px;vertical-align:top;white-space:pre}}
  .diff-table td.ln{{width:18px;text-align:center;color:var(--muted);
    font-size:9px;user-select:none;border-right:1px solid var(--border)}}
  .diff-table pre{{font-family:inherit;font-size:11px;white-space:pre-wrap;
    word-break:break-all}}
</style>
</head>
<body>
<div class="layout">
  <nav class="sidebar">
    <h2>Sections ({len(rows)})</h2>
    <ul>{toc_items}</ul>
  </nav>
  <main class="main">
    <div class="page-header">
      <h1>PDF Comparison Report</h1>
      <div class="meta-grid">
        <div class="meta-box">
          <div class="lbl">File A</div>
          <div class="val" style="font-size:13px">{esc(name_a)}</div>
          <div class="sub">{esc(meta_a.get('extracted_at',''))}</div>
        </div>
        <div class="meta-box">
          <div class="lbl">File B</div>
          <div class="val" style="font-size:13px">{esc(name_b)}</div>
          <div class="sub">{esc(meta_b.get('extracted_at',''))}</div>
        </div>
        <div class="meta-box">
          <div class="lbl">Overall Similarity</div>
          <div class="val" style="color:{_sim_colour(avg_sim)}">{avg_sim}%</div>
          <div class="sub">avg across {matched_count} matched sections</div>
        </div>
        <div class="meta-box">
          <div class="lbl">Matched / Missing</div>
          <div class="val">{matched_count}</div>
          <div class="sub">{only_a_count} only in A · {only_b_count} only in B</div>
        </div>
        <div class="meta-box">
          <div class="lbl">Generated</div>
          <div class="val" style="font-size:13px">{datetime.now().strftime('%Y-%m-%d')}</div>
          <div class="sub">{datetime.now().strftime('%H:%M:%S')}</div>
        </div>
      </div>
    </div>

    <div class="section-heading">All Sections — {len(rows)} total</div>
    {"".join(cards)}
  </main>
</div>
</body>
</html>"""

    path = out_dir / "_COMPARISON_REPORT.html"
    path.write_text(html, encoding="utf-8")
    print(f"  ✓  _COMPARISON_REPORT.html  ({len(rows)} sections)")
    return path


# ══════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════

def write_excel_report(rows: list, name_a: str, name_b: str,
                       out_dir: Path,
                       meta_a: dict = None, meta_b: dict = None) -> Path:
    """
    Write a four-sheet Excel workbook.

      Sheet 1 — Summary
      Sheet 2 — All Sections   (every row, matched + missing)
      Sheet 3 — Unmatched      (only-in-A and only-in-B rows)
      Sheet 4 — Section Detail (full content side-by-side)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠  openpyxl not installed — skipping Excel. Run: pip install openpyxl")
        return None

    meta_a = meta_a or {}
    meta_b = meta_b or {}
    wb = Workbook()

    # ── Shared style helpers ──────────────────────────────────────
    FN = "Arial"
    thin  = Side(style="thin", color="CBD5E1")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _font(size=10, bold=False, color="000000", italic=False):
        return Font(name=FN, size=size, bold=bold, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _align(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="top", wrap_text=wrap)

    def _hdr_cell(ws, row, col, value,
                  bg="1E3A5F", fg="FFFFFF", size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(size=size, bold=True, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = bdr
        return c

    def _data_cell(ws, row, col, value, bg=None, wrap=False, bold=False, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = _font(bold=bold, italic=italic, color="000000")
        c.fill      = _fill(bg) if bg else PatternFill()
        c.alignment = _align(wrap=wrap)
        c.border    = bdr
        return c

    def _set_widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    STATUS_FILLS = {
        "CHANGED":    "FEF3C7",
        "IDENTICAL":  "D1FAE5",
        "ONLY IN A":  "DBEAFE",
        "ONLY IN B":  "DCFCE7",
    }
    MISSING_FILL = "E5E7EB"

    def _sim_fill(pct):
        if pct >= 90: return "D1FAE5"
        if pct >= 70: return "FEF3C7"
        if pct >= 40: return "FFEDD5"
        return "FEE2E2"

    # ════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "PDF Comparison Report"
    c.font  = _font(size=16, bold=True, color="FFFFFF")
    c.fill  = _fill("1D4ED8")
    c.alignment = _align("center")
    ws1.row_dimensions[1].height = 30

    meta_rows = [
        ("File A",     name_a),
        ("File B",     name_b),
        ("Extracted A",meta_a.get("extracted_at","—")),
        ("Extracted B",meta_b.get("extracted_at","—")),
        ("Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for r, (lbl, val) in enumerate(meta_rows, 2):
        _hdr_cell(ws1, r, 1, lbl)
        ws1.merge_cells(f"B{r}:D{r}")
        _data_cell(ws1, r, 2, val)

    ws1.row_dimensions[7].height = 12  # gap

    matched_count  = sum(1 for r in rows if r["status"] in ("CHANGED","IDENTICAL"))
    identical_count= sum(1 for r in rows if r["status"] == "IDENTICAL")
    only_a_count   = sum(1 for r in rows if r["status"] == "ONLY IN A")
    only_b_count   = sum(1 for r in rows if r["status"] == "ONLY IN B")
    sims = [r["stats"]["line_similarity_pct"]
            for r in rows if r["status"] in ("CHANGED","IDENTICAL")]
    avg_sim = round(sum(sims)/len(sims),1) if sims else 0

    _hdr_cell(ws1, 8, 1, "Metric")
    _hdr_cell(ws1, 8, 2, "Value")
    ws1.merge_cells("C8:D8"); _hdr_cell(ws1, 8, 3, "Notes")

    stat_rows = [
        ("Total sections",    len(rows),         "All sections across both documents"),
        ("Matched",           matched_count,      f"{identical_count} identical, {matched_count-identical_count} differ"),
        ("Only in A",         only_a_count,       f"Missing from {name_b}"),
        ("Only in B",         only_b_count,       f"Missing from {name_a}"),
        ("Avg similarity",    f"{avg_sim}%",      "Across matched section pairs"),
    ]
    for i, (m, v, d) in enumerate(stat_rows):
        r = 9 + i
        bg = "F0F4FF" if i % 2 == 0 else None
        _data_cell(ws1, r, 1, m, bg=bg, bold=True)
        _data_cell(ws1, r, 2, v, bg=bg)
        ws1.merge_cells(f"C{r}:D{r}")
        _data_cell(ws1, r, 3, d, bg=bg)

    _set_widths(ws1, {"A":22,"B":18,"C":45,"D":10})

    # ════════════════════════════════════
    # Sheet 2 — All Sections
    # ════════════════════════════════════
    ws2 = wb.create_sheet("All Sections")
    hdrs = ["#","Status","Level",
            f"Title — {name_a}", f"Title — {name_b}",
            "Title Match %","Content Sim %",
            "Lines A","Lines B","+Added","-Removed","=Unchanged",
            "AI Summary"]
    for c, h in enumerate(hdrs, 1):
        _hdr_cell(ws2, 1, c, h)
    ws2.row_dimensions[1].height = 20

    for i, row in enumerate(rows, 1):
        r     = i + 1
        st    = row["status"]
        stats = row["stats"]
        sim   = stats["line_similarity_pct"]
        bg    = STATUS_FILLS.get(st, "FFFFFF")

        # Title cells — grey out the missing one
        ta_val = row["title_a"]
        tb_val = row["title_b"]
        ta_bg  = MISSING_FILL if ta_val == MISSING_PLACEHOLDER else bg
        tb_bg  = MISSING_FILL if tb_val == MISSING_PLACEHOLDER else bg
        sim_bg = _sim_fill(sim) if st in ("CHANGED","IDENTICAL") else MISSING_FILL

        vals = [i, st, row["level"].capitalize(), ta_val, tb_val,
                round(row["title_score"],1) if row["title_score"] else "",
                sim if st in ("CHANGED","IDENTICAL") else "",
                stats["total_lines_a"] or "",
                stats["total_lines_b"] or "",
                stats["lines_added"]   or "",
                stats["lines_removed"] or "",
                stats["lines_unchanged"] or "",
                row["summary"] or ""]

        for c, val in enumerate(vals, 1):
            cell_bg = (ta_bg if c==4 else tb_bg if c==5
                       else sim_bg if c==7 else bg)
            is_missing_ta = (c == 4 and ta_val == MISSING_PLACEHOLDER)
            is_missing_tb = (c == 5 and tb_val == MISSING_PLACEHOLDER)
            italic = is_missing_ta or is_missing_tb
            _data_cell(ws2, r, c, val, bg=cell_bg, wrap=(c==13), italic=italic)

        ws2.row_dimensions[r].height = 55 if row["summary"] else 18

    # Totals row
    tr = len(rows) + 2
    _data_cell(ws2, tr, 1, "TOTALS", bold=True)
    for col_idx in range(8, 13):
        c = ws2.cell(row=tr, column=col_idx,
                     value=f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{tr-1})")
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("1E3A5F")
        c.border = bdr
        c.alignment = _align("center")

    _set_widths(ws2, {"A":4,"B":12,"C":11,"D":32,"E":32,
                      "F":12,"G":12,"H":9,"I":9,"J":9,"K":9,"L":11,"M":55})
    ws2.freeze_panes = "A2"

    # ════════════════════════════════════
    # Sheet 3 — Unmatched
    # ════════════════════════════════════
    ws3 = wb.create_sheet("Unmatched")
    for c, h in enumerate(["Source","Level","Title","Content Preview"], 1):
        _hdr_cell(ws3, 1, c, h)
    ws3.row_dimensions[1].height = 18

    r = 2
    for row in rows:
        st = row["status"]
        if st not in ("ONLY IN A", "ONLY IN B"):
            continue
        if st == "ONLY IN A":
            source  = f"Only in A — {name_a}"
            title   = row["title_a"]
            content = row["content_a"]
            bg      = STATUS_FILLS["ONLY IN A"]
        else:
            source  = f"Only in B — {name_b}"
            title   = row["title_b"]
            content = row["content_b"]
            bg      = STATUS_FILLS["ONLY IN B"]

        preview = content[:300].replace("\n", " ")
        for c, val in enumerate([source, row["level"].capitalize(), title, preview], 1):
            _data_cell(ws3, r, c, val, bg=bg, wrap=(c==4))
        ws3.row_dimensions[r].height = 38
        r += 1

    if r == 2:
        ws3.cell(row=2, column=1).value = "All sections were matched — no missing sections."
        ws3.cell(row=2, column=1).font  = _font(italic=True, color="6B7280")

    _set_widths(ws3, {"A":30,"B":12,"C":45,"D":70})
    ws3.freeze_panes = "A2"

    # ════════════════════════════════════
    # Sheet 4 — Section Detail
    # ════════════════════════════════════
    ws4 = wb.create_sheet("Section Detail")
    for c, h in enumerate(["#","Status","Level","Title","Sim %",
                            f"Content — {name_a}",
                            f"Content — {name_b}"], 1):
        _hdr_cell(ws4, 1, c, h)
    ws4.row_dimensions[1].height = 20

    for i, row in enumerate(rows, 1):
        r      = i + 1
        st     = row["status"]
        sim    = row["stats"]["line_similarity_pct"]
        bg     = STATUS_FILLS.get(st, "FFFFFF")
        sim_bg = _sim_fill(sim) if st in ("CHANGED","IDENTICAL") else MISSING_FILL

        title  = (row["title_a"] if st != "ONLY IN B" else row["title_b"])

        ca_val = row["content_a"] if row["content_a"] else MISSING_PLACEHOLDER
        cb_val = row["content_b"] if row["content_b"] else MISSING_PLACEHOLDER
        ca_bg  = MISSING_FILL if not row["content_a"] else bg
        cb_bg  = MISSING_FILL if not row["content_b"] else bg
        ca_it  = not bool(row["content_a"])
        cb_it  = not bool(row["content_b"])

        _data_cell(ws4, r, 1, i,        bg=bg)
        _data_cell(ws4, r, 2, st,       bg=bg, bold=True)
        _data_cell(ws4, r, 3, row["level"].capitalize(), bg=bg)
        _data_cell(ws4, r, 4, title,    bg=bg)
        _data_cell(ws4, r, 5,
                   sim if st in ("CHANGED","IDENTICAL") else "",
                   bg=sim_bg)
        _data_cell(ws4, r, 6, ca_val, bg=ca_bg, wrap=True, italic=ca_it)
        _data_cell(ws4, r, 7, cb_val, bg=cb_bg, wrap=True, italic=cb_it)

        max_lines = max(
            len(row["content_a"].splitlines()) if row["content_a"] else 1,
            len(row["content_b"].splitlines()) if row["content_b"] else 1,
        )
        ws4.row_dimensions[r].height = min(max_lines * 14, 400)

    _set_widths(ws4, {"A":4,"B":12,"C":11,"D":38,"E":9,"F":70,"G":70})
    ws4.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────
    path = out_dir / "_COMPARISON_REPORT.xlsx"
    wb.save(str(path))
    matched_ct = sum(1 for r in rows if r["status"] in ("CHANGED","IDENTICAL"))
    missing_ct = sum(1 for r in rows if r["status"] in ("ONLY IN A","ONLY IN B"))
    print(f"  ✓  _COMPARISON_REPORT.xlsx  "
          f"({matched_ct} matched, {missing_ct} missing sections)")
    return path


# ══════════════════════════════════════════════════════════════════
# JSON METADATA
# ══════════════════════════════════════════════════════════════════

def write_json_metadata(rows: list, name_a: str, name_b: str,
                        out_dir: Path) -> None:
    matched = [r for r in rows if r["status"] in ("CHANGED","IDENTICAL")]
    sims    = [r["stats"]["line_similarity_pct"] for r in matched]
    data = {
        "file_a":    name_a,
        "file_b":    name_b,
        "generated": datetime.now().isoformat(),
        "summary": {
            "total_sections": len(rows),
            "matched":        len(matched),
            "only_in_a":      sum(1 for r in rows if r["status"]=="ONLY IN A"),
            "only_in_b":      sum(1 for r in rows if r["status"]=="ONLY IN B"),
            "avg_similarity": round(sum(sims)/len(sims),1) if sims else 0,
        },
        "sections": [
            {
                "status":       r["status"],
                "level":        r["level"],
                "title_a":      r["title_a"],
                "title_b":      r["title_b"],
                "title_score":  r["title_score"],
                "ai_summary":   r["summary"],
                **r["stats"],
            }
            for r in rows
        ],
    }
    path = out_dir / "_MATCHED.json"
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"  ✓  _MATCHED.json")


# ══════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════

def parse_args():
    parser = argparse.ArgumentParser(
        description="Compare two pdf_section_extractor output directories",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Workflow:
  # Step 1 — extract each PDF
  python pdf_section_extractor.py doc_a.pdf extracted_a/
  python pdf_section_extractor.py doc_b.pdf extracted_b/

  # (Optional) Step 2 — generate image descriptions
  python pdf_image_describer.py extracted_a/
  python pdf_image_describer.py extracted_b/

  # Step 3 — compare
  python pdf_compare.py extracted_a/ extracted_b/

  # Re-run comparison with different options (no re-extraction needed):
  python pdf_compare.py extracted_a/ extracted_b/ --skip-summary
  python pdf_compare.py extracted_a/ extracted_b/ --fuzzy-threshold 60
  python pdf_compare.py extracted_a/ extracted_b/ --text-model mistral

Ollama text models for summaries:
  llama3.2   ollama pull llama3.2
  mistral    ollama pull mistral
  llama3     ollama pull llama3
  phi3       ollama pull phi3
        """,
    )
    parser.add_argument("dir_a", help="Extractor output directory for the first PDF")
    parser.add_argument("dir_b", help="Extractor output directory for the second PDF")
    parser.add_argument("output_dir", nargs="?", default="pdf_comparison",
                        help="Where to write comparison outputs (default: pdf_comparison/)")
    parser.add_argument("--text-model",      default=DEFAULT_TEXT_MODEL,
                        help=f"Ollama text model for summaries (default: {DEFAULT_TEXT_MODEL})")
    parser.add_argument("--ollama-url",      default=DEFAULT_OLLAMA_URL)
    parser.add_argument("--skip-summary",    action="store_true",
                        help="Skip AI summaries (faster)")
    parser.add_argument("--fuzzy-threshold", type=float, default=75.0,
                        help="Min title similarity %% for a match (default: 75)")
    parser.add_argument("--report-format",   choices=["html","text","both"],
                        default="html",
                        help="html | text | both  (Excel always written, default: html)")
    return parser.parse_args()


def main():
    args    = parse_args()
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'═'*62}")
    print(f"  PDF Comparator")
    print(f"{'═'*62}")
    print(f"  Source A      : {args.dir_a}/")
    print(f"  Source B      : {args.dir_b}/")
    print(f"  Output        : {args.output_dir}/")
    print(f"  Match threshold: {args.fuzzy_threshold}%")
    print(f"  Text model    : {args.text_model}")
    print(f"  AI summaries  : {'disabled' if args.skip_summary else 'enabled'}")
    print(f"{'═'*62}\n")

    # ── Step 1/3: Load extractor output directories ───────────────
    print("Step 1/3  Loading extractor directories...\n")

    sections_a, catalog_a, meta_a, state_a = load_extraction_dir(args.dir_a, "A")
    name_a = meta_a.get("source_pdf") or Path(args.dir_a).name
    print(_dir_summary_str("A", args.dir_a, sections_a, catalog_a, meta_a))
    print()

    sections_b, catalog_b, meta_b, state_b = load_extraction_dir(args.dir_b, "B")
    name_b = meta_b.get("source_pdf") or Path(args.dir_b).name
    print(_dir_summary_str("B", args.dir_b, sections_b, catalog_b, meta_b))
    print()

    # ── Step 2/3: Match, summarise, build unified row list ────────
    print(f"Step 2/3  Matching sections (threshold: {args.fuzzy_threshold}%)...\n")

    matched_raw, only_a, only_b = match_sections(
        sections_a, sections_b, threshold=args.fuzzy_threshold
    )
    print(f"  Matched  : {len(matched_raw)}")
    print(f"  Only A   : {len(only_a)}")
    print(f"  Only B   : {len(only_b)}\n")

    print(f"  Generating AI summaries with '{args.text_model}'...\n")
    matched = summarise_all_matches(
        matched_raw,
        name_a       = name_a,
        name_b       = name_b,
        base_url     = args.ollama_url,
        model        = args.text_model,
        skip_summary = args.skip_summary,
    )

    # Build single unified row list (matched + only_a + only_b)
    rows = build_all_rows(matched, only_a, only_b)

    # ── Step 3/3: Write outputs ───────────────────────────────────
    print(f"\nStep 3/3  Writing outputs to {args.output_dir}/\n")

    write_html_report(rows, name_a, name_b, out_dir, meta_a, meta_b)
    write_excel_report(rows, name_a, name_b, out_dir, meta_a, meta_b)
    write_json_metadata(rows, name_a, name_b, out_dir)

    if args.report_format in ("text", "both"):
        _write_text_report(rows, name_a, name_b, out_dir)

    # ── Summary ───────────────────────────────────────────────────
    sims = [r["stats"]["line_similarity_pct"]
            for r in rows if r["status"] in ("CHANGED","IDENTICAL")]
    avg  = round(sum(sims)/len(sims),1) if sims else 0
    identical = sum(1 for r in rows if r["status"]=="IDENTICAL")

    print(f"\n{'═'*62}")
    print(f"  ✅  Comparison complete!")
    print(f"      Total sections   : {len(rows)}")
    print(f"      Matched          : {len(matched)}  ({identical} identical)")
    print(f"      Only in A        : {len(only_a)}")
    print(f"      Only in B        : {len(only_b)}")
    print(f"      Avg similarity   : {avg}%")
    print(f"      HTML report      : {args.output_dir}/_COMPARISON_REPORT.html")
    print(f"      Excel workbook   : {args.output_dir}/_COMPARISON_REPORT.xlsx")
    print(f"{'═'*62}\n")


def _write_text_report(rows: list, name_a: str, name_b: str, out_dir: Path) -> None:
    """Optional plain-text fallback report."""
    lines = [
        "PDF COMPARISON REPORT", "="*70,
        f"  File A    : {name_a}",
        f"  File B    : {name_b}",
        f"  Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "="*70, "",
    ]
    for i, row in enumerate(rows, 1):
        st  = row["status"]
        sim = row["stats"]["line_similarity_pct"]
        lines.append(f"\n[{i}] {st}  [{row['level'].upper()}]")
        if st in ("CHANGED","IDENTICAL"):
            lines.append(f"  Title A  : {row['title_a']}")
            lines.append(f"  Title B  : {row['title_b']}")
            lines.append(f"  Similarity: {sim}%")
            if row.get("summary"):
                wrapped = "\n".join("  "+l for l in textwrap.wrap(row["summary"],65))
                lines += ["  AI Summary:", wrapped]
        elif st == "ONLY IN A":
            lines.append(f"  Title : {row['title_a']}")
            lines.append(f"  ⚠  Not present in {name_b}")
        else:
            lines.append(f"  Title : {row['title_b']}")
            lines.append(f"  ⚠  Not present in {name_a}")
    path = out_dir / "_COMPARISON_REPORT.txt"
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  ✓  _COMPARISON_REPORT.txt")


if __name__ == "__main__":
    main()
